#taking the information from a colum A and putting into a txt file
#--------------------------------------------------------------------
from openpyxl import load_workbook

wb = load_workbook(filename = 'exoplanet.eu_catalog.xlsx')
ws = wb.active
file = open('27-05-19-16.txt', 'a')

cell_number = len(ws['B'])

counter = 0
for i in ws['B']:
     counter += 1
     current_cell_name = f'A{counter}'m 
     if counter == cell_number:
          file.write(f"{ws[current_cell_name].value}")
     else:
          file.write(f"{ws[current_cell_name].value}\n")
     print(counter)

print(f'CELL NUM: {cell_number}')
print(f'COUNTER: {counter}')

file.close()
wb.close
#--------------------------------------------------------------------



#filling up the empety spaces
#---------------------------------------------------
# from openpyxl import load_workbook
# import openpyxl.utils.cell

# wb = load_workbook('exoplanet.eu_catalog.xlsx')
# ws = wb.active

# columns = []

# for i in ws[1]:
#      letter = openpyxl.utils.cell.get_column_letter(i.column)
#      columns.append(letter)
#      print(i.value)
# print(columns)

# for column in columns:
#      for cell in ws[column]:
#           if cell.value == None:
#                cell.value = '-'

# wb.save('test.xlsx')
# wb.close()

#procurando um informação em uma célula
#---------------------------------------------------
# import openpyxl
# from openpyxl import load_workbook
# import openpyxl.utils.cell

# print('OPENING PLANETS DATABASE...')
# wb = load_workbook('exoplanet.eu_catalog.xlsx')
# ws = wb.active

# target = 'GLASS-JWST-BD1'
# print(f'LOOKING FOR TARGET: "{target}"')

# for i in ws['A']:
#     if i.value == target:
#          column = openpyxl.utils.cell.get_column_letter(i.column)
#          row = i.row
#          print('TARGET FOUND.')
#          print(f'COlUMN: {column}')
#          print(f'ROW: {row}')


# wb.close